from io import BytesIO

from flask import (
    send_file,
    session
)

from openpyxl import Workbook

from database import app

from models import (
    Usuario,
    Trabalho,
    Avaliacao
)


@app.route(
    '/exportar-planilha'
)
def exportar_planilha():

    if 'usuario_id' not in session:
        return "Acesso negado!"

    if session['tipo'] != 'admin':
        return "Somente administradores podem exportar dados!"

    usuarios = Usuario.query.all()

    trabalhos = Trabalho.query.all()

    avaliacoes = Avaliacao.query.all()

    workbook = Workbook()

    # ==========================
    # ABA USUÁRIOS
    # ==========================

    sheet_usuarios = workbook.active

    sheet_usuarios.title = "Usuários"

    sheet_usuarios.append([
        "Nome",
        "Email",
        "Tipo",
        "Matricula",
        "Ativo"
    ])

    for usuario in usuarios:

        sheet_usuarios.append([
            usuario.nome,
            usuario.email,
            usuario.tipo_usuario,
            usuario.matricula,
            usuario.ativo
        ])

    # ==========================
    # ABA TRABALHOS
    # ==========================

    sheet_trabalhos = (
        workbook.create_sheet(
            "Trabalhos"
        )
    )

    sheet_trabalhos.append([
        "Titulo",
        "Status",
        "Media Final",
        "Aluno ID",
        "Criado Em"
    ])

    for trabalho in trabalhos:

        sheet_trabalhos.append([
            trabalho.titulo,
            trabalho.status,
            trabalho.media_final,
            trabalho.aluno_id,
            trabalho.criado_em
        ])

    # ==========================
    # ABA AVALIAÇÕES
    # ==========================

    sheet_avaliacoes = (
        workbook.create_sheet(
            "Avaliacoes"
        )
    )

    sheet_avaliacoes.append([
        "Trabalho ID",
        "Avaliador ID",
        "Originalidade",
        "Clareza",
        "Metodologia",
        "Relevancia",
        "Organizacao",
        "Comentario"
    ])

    for avaliacao in avaliacoes:

        sheet_avaliacoes.append([
            avaliacao.trabalho_id,
            avaliacao.avaliador_id,
            avaliacao.originalidade,
            avaliacao.clareza,
            avaliacao.metodologia,
            avaliacao.relevancia,
            avaliacao.organizacao,
            avaliacao.comentario
        ])

    arquivo = BytesIO()

    workbook.save(
        arquivo
    )

    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=
        "dados_sistema.xlsx",
        mimetype=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )